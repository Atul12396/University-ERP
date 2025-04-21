import datetime
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CustomUser, Subject, Attendance, AttendanceReport, Student, Teacher, Marks, Course, Timetable, Marks, Branch, Semester, Announcement, FeeStructure, Payment,FinancialFees,TransportFees,Transport,Leave,HOD
from django.http import HttpResponse
from django.http import HttpResponseForbidden
from django.shortcuts import render, get_object_or_404
from .forms import SubjectForm, StudentEditForm, StudentForm, BranchForm, SemesterForm
from django.utils import timezone
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from .forms import MarksUploadForm
from django.db import transaction
from datetime import datetime as dt
from registration_app.forms import TimetableForm, PaymentForm
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from django.db.models import Q
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import Course, Branch, Semester, Student
from .forms import StudentUpdateForm




def view_courses(request):
    courses = Course.objects.all()
    return render(request, 'view_courses.html', {'courses': courses})

def view_branches(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    branches = Branch.objects.filter(course=course)
    
    if branches.exists():
        return render(request, 'view_branches.html', {'course': course, 'branches': branches})
    else:
        return redirect('view_semesters_no_branch', course_id=course_id)

def view_semesters(request, course_id, branch_id=None):
    course = get_object_or_404(Course, pk=course_id)
    branch = None
    
    if branch_id:
        branch = get_object_or_404(Branch, pk=branch_id)
        semesters = Semester.objects.filter(course=course, branch=branch)
    else:
        semesters = Semester.objects.filter(course=course)
    
    return render(request, 'view_semesters.html', {'course': course, 'branch': branch, 'semesters': semesters})

def view_semesters_no_branch(request, course_id):
    course = get_object_or_404(Course, pk=course_id)
    semesters = Semester.objects.filter(course=course)
    return render(request, 'view_semesters_no_branch.html', {'course': course, 'semesters': semesters})

def view_subjects(request, course_id, semester_id, branch_id=None):
    course = get_object_or_404(Course, pk=course_id)
    semester = get_object_or_404(Semester, pk=semester_id)
    
    if branch_id is not None:
        branch = get_object_or_404(Branch, pk=branch_id)
        subjects = Subject.objects.filter(course=course, semester=semester, branches=branch)
    else:
        # If branch_id is not provided, fetch subjects only filtered by course and semester
        subjects = Subject.objects.filter(course=course, semester=semester)
    
    return render(request, 'view_subjects.html', {'subjects': subjects})

def view_subjects(request, course_id, semester_id, branch_id):
    course = get_object_or_404(Course, pk=course_id)
    semester = get_object_or_404(Semester, pk=semester_id)
    branch = get_object_or_404(Branch, pk=branch_id)
    subjects = Subject.objects.filter(course=course, semester=semester, branches=branch)
    return render(request, 'view_subjects.html', {'subjects': subjects, 'branch': branch, 'course': course, 'semester': semester})


def view_subjects_no_branch(request, course_id, semester_id):
    course = get_object_or_404(Course, pk=course_id)
    semester = get_object_or_404(Semester, pk=semester_id)
    subjects = Subject.objects.filter(course=course, semester=semester)
    return render(request, 'view_subjects_no_branch.html', {'subjects': subjects, 'course': course, 'semester': semester})

def parse_time_string(time_str):
    # Convert time string to 24-hour format
    time_obj = datetime.strptime(time_str, '%I:%M %p').time()
    # Convert time object to Django's timezone-aware time
    return timezone.make_aware(datetime.combine(timezone.now().date(), time_obj))

def create_timetable_no_branch(request, course_id, semester_id, subject_id):
    if request.method == 'POST':
        day = request.POST['day']
        start_time_str = request.POST['start_time']
        end_time_str = request.POST['end_time']
        venue = request.POST['venue']
        
        # Parse time strings into datetime objects
        start_time = datetime.strptime(start_time_str, '%I:%M %p').time()
        end_time = datetime.strptime(end_time_str, '%I:%M %p').time()
        
        timetable = Timetable.objects.create(
            course_id=course_id,
            semester_id=semester_id,
            subject_id=subject_id,
            day=day,
            start_time=start_time,
            end_time=end_time,
            venue=venue
        )
        return redirect('timetable_success')
    else:
        return render(request, 'create_timetable_no_branch.html', {'course_id': course_id, 'semester_id': semester_id, 'subject_id': subject_id})


def create_timetable_with_branch(request, course_id, branch_id, semester_id, subject_id):
    if request.method == 'POST':
        day = request.POST['day']
        start_time_str = request.POST['start_time']
        end_time_str = request.POST['end_time']
        venue = request.POST['venue']
        
        # Parse time strings into datetime objects
        start_time = datetime.strptime(start_time_str, '%I:%M %p').time()
        end_time = datetime.strptime(end_time_str, '%I:%M %p').time()
        
        timetable = Timetable.objects.create(
            course_id=course_id,
            branch_id=branch_id,
            semester_id=semester_id,
            subject_id=subject_id,
            day=day,
            start_time=start_time,
            end_time=end_time,
            venue=venue
        )
        return redirect('timetable_success')
    else:
        return render(request, 'create_timetable_with_branch.html', {'course_id': course_id, 'branch_id': branch_id, 'semester_id': semester_id, 'subject_id': subject_id})

def timetable_success(request):
    return HttpResponse("Timetable created successfully!")
    


@login_required
def view_timetable(request):
    user = request.user
    if hasattr(user, 'teacher'):
        subjects = user.teacher.subjects.all()
        timetable_entries = Timetable.objects.filter(subject__in=subjects)
    elif hasattr(user, 'student'):
        subjects = user.student.subjects.all()
        timetable_entries = Timetable.objects.filter(subject__in=subjects)
    else:
        timetable_entries = None
    return render(request, 'timetable.html', {'timetable_entries': timetable_entries})


@login_required(login_url='login')
def HomePage(request):
    return render(request, 'home.html')

def SignupPage(request):
    if request.method == 'POST':
        uname = request.POST.get('username')
        email = request.POST.get('email')
        pass1 = request.POST.get('password1')
        pass2 = request.POST.get('password2')

        if pass1 != pass2:
             return HttpResponse("Your password and confirm password are not the same!!")
        else:
            # Check if the username already exists
            if CustomUser.objects.filter(username=uname).exists():
                messages.error(request, "Username already exists. Please choose a different one.")
            else:
                # Extract the role from the email address
                if 'student' in email:
                    role = 'student'
                elif 'teacher' in email:
                    role = 'teacher'
                elif 'hod' in email:
                    role = 'hod'
                elif 'ss' in email:
                    role = 'ss'
                elif 'as' in email:
                    role = 'as'
                else:
                    role = ''  # Handle an invalid email format

                if role:
                    # Create a user with the extracted role
                    my_user = CustomUser.objects.create_user(username=uname, email=email, password=pass1, role=role)
                    my_user.save()
                    messages.success(request, "User created successfully.")
                    return redirect('login')
                else:
                    messages.error(request, "Invalid email format. Please use 'name.role@university.org'.")

    return render(request, 'signup.html')

def custom_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.role == 'student':
                return redirect('student_dashboard')
            elif user.role == 'teacher':
                return redirect('teacher_dashboard')
            elif user.role == 'hod':
                return redirect('hod-dashboard')
            elif user.role == 'ss':
                return redirect('student_section_dashboard')
        else:
            messages.error(request, "Username or password is incorrect.")

    return render(request, 'login.html')

from django.contrib import messages



@login_required(login_url='login')
def student_dashboard(request):
    try:
        student = request.user.student
        student_details = {
            'name': student.name,
            'roll_number': student.roll_number,
            'student_id': student.student_id,
            'branch': student.branch,
            'semester': student.semester,
            'contact': student.contact,
            'email': student.user.email,
            'address': student.address,
        }

        # Fetch fee structure for the student
        fee_structure = student.fee_structure
        fee_amount = (
            fee_structure.registration_fees + fee_structure.academic_fees
            if fee_structure else None
        )

        fee_status = student.payment_status

        # Check payment status
        if not fee_status:
            messages.error(request, "Please complete your payment to access the dashboard.")
            return redirect('pay_fees')  # Redirect to payment page if payment is pending

    except Student.DoesNotExist:
        return HttpResponseForbidden("You don't have permission to access this page.")

    subjects = Subject.objects.filter(students=student)

    subject_data = []
    for subject in subjects:
        total_classes = subject.total_classes
        attendance_reports = AttendanceReport.objects.filter(student=student, attendance__subject=subject)
        attended_classes = attendance_reports.filter(status=True).count()
        attendance_percentage = (attended_classes / total_classes) * 100 if total_classes > 0 else 0

        subject_data.append({
            'subject_name': subject.name,
            'attendance_percentage': attendance_percentage,
        })

    # Retrieve latest announcements
    announcements = Announcement.objects.all().order_by('-posted_on')[:3]

    return render(request, 'index.html', {
        'subject_data': subject_data,
        'student_details': student_details,
        'announcements': announcements,
        'fee_amount': fee_amount,
        'fee_status': fee_status,
    })






@login_required(login_url='login')
def hod_dashboard(request):
    if request.user.role == 'hod':
        hod_department = request.user.department
        name=request.user.hod

        # Fetch department-specific data
        courses = Course.objects.filter(Q(department=hod_department) | Q(department=None))
        subjects = Subject.objects.filter(Q(course__department=hod_department) | Q(course__department=None))
        teachers = Teacher.objects.filter(Q(branch__course__department=hod_department) | Q(branch__course__department=None))
        students = Student.objects.filter(Q(branch__course__department=hod_department) | Q(branch__course__department=None))
        semesters = Semester.objects.filter(Q(course__department=hod_department) | Q(course__department=None))
        branches = Branch.objects.filter(Q(course__department=hod_department) | Q(course__department=None))

        # Fetch pending leave requests
        pending_leaves = Leave.objects.filter(teacher__department=hod_department, status='Pending')

        context = {
            'courses': courses,
            'subjects': subjects,
            'teachers': teachers,
            'students': students,
            'semesters': semesters,
            'branches': branches,
            'pending_leaves': pending_leaves,
            'name':name,
        }

        return render(request, 'hod/hod-dashboard.html', context)

    return HttpResponseForbidden("You don't have permission to access this page.")

@login_required(login_url='login')
def hod_leave_requests(request):
    """View all pending leave requests for the HOD's department"""
    if request.user.role != 'hod':
        return HttpResponseForbidden("You don't have permission to access this page.")

    # Ensure HOD object is fetched properly
    try:
        hod = HOD.objects.get(user=request.user)
    except HOD.DoesNotExist:
        return HttpResponseForbidden("HOD profile not found.")

    # Get pending leaves for teachers in the HOD's department
    pending_leaves = Leave.objects.filter(teacher__department=hod.department, status='Pending')

    return render(request, 'hod/leave_requests.html', {'pending_leaves': pending_leaves})


 
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required(login_url='login')
def manage_leave_request(request, leave_id, action):
    """Approve or reject leave requests via AJAX"""
    if request.user.role != 'hod':
        return JsonResponse({'success': False, 'error': "Unauthorized access"})

    try:
        leave = Leave.objects.get(id=leave_id)
        if action == "approve":
            leave.status = "Approved"
        elif action == "reject":
            leave.status = "Rejected"
        else:
            return JsonResponse({'success': False, 'error': "Invalid action"})

        leave.save()

        return JsonResponse({'success': True})
    except Leave.DoesNotExist:
        return JsonResponse({'success': False, 'error': "Leave request not found"})



from django.shortcuts import render
from .models import Student

def hod_students(request):
    if request.user.role == 'hod':
        # Retrieve students belonging to the HOD's department or course
        hod_department = request.user.hod.department
        students = Student.objects.filter(course__department=hod_department)
        
        context = {
            'students': students,
        }

        return render(request, 'hod/students.html', context)
    
    # If the user is not an HOD, handle accordingly
    return HttpResponseForbidden("You don't have permission to access this page.")





@login_required(login_url='login')
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('hod_students')  # Redirect to the list of students
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'hod/edit_student.html', {'form': form, 'student': student})

def update_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('hod_dashboard')
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'hod/edit_student.html', {'form': form, 'student': student})



from django.shortcuts import render
from .models import Branch

def hod_branches(request):
    if request.user.role == 'hod':
        # Retrieve branches belonging to the HOD's department
        hod_department = request.user.hod.department
        branches = Branch.objects.filter(department=hod_department)
        
        context = {
            'branches': branches,
        }

        return render(request, 'hod/branches.html', context)
    
    # If the user is not an HOD, handle accordingly
    return HttpResponseForbidden("You don't have permission to access this page.")



from django.shortcuts import render
from .models import Teacher

def hod_teachers(request):
    if request.user.role == 'hod':
        # Retrieve teachers assigned to the HOD's department
        hod_department = request.user.hod.department
        teachers = Teacher.objects.filter(department=hod_department)
        
        context = {
            'teachers': teachers,
        }

        return render(request, 'hod/hod-teachers.html', context)
    
    # If the user is not an HOD, handle accordingly
    return HttpResponseForbidden("You don't have permission to access this page.")



@login_required(login_url='login')
def edit_branch(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    if request.method == 'POST':
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            return redirect('hod-dashboard')
    else:
        form = BranchForm(instance=branch)
    return render(request, 'hod/edit_branch.html', {'form': form})

 

@login_required(login_url='login')
def hod_subjects(request):
    if request.user.role == 'hod':
        # Retrieve the HOD's department
        hod_department = request.user.hod.department
        
        # Filter subjects based on the department of the HOD
        subjects = Subject.objects.filter(course__department=hod_department)
        
        context = {
            'subjects': subjects,
        }

        return render(request, 'hod/subjects.html', context)
    
    return HttpResponseForbidden("You don't have permission to access this page.")


# views.py

from django.shortcuts import render
from .forms import SubjectForm
from .models import CustomUser, Course, Department

def add_subject(request):
    hod = request.user.hod
    department_id = None
    if 'department' in request.GET:
        department_id = request.GET['department']

    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to a success page or do something else
    else:
        form = SubjectForm()

        # Filter teachers based on their role (teacher) and courses based on the department
        teachers = CustomUser.objects.filter(role='teacher', teacher__department=hod.department)
        courses = Course.objects.filter(department_id=department_id) if department_id else Course.objects.none()
        form.fields['teacher'].queryset = teachers

        # Filter semesters based on the department of the logged-in HOD
        semesters = Semester.objects.filter(branch__department=hod.department)

        # Pass the filtered queryset to the form
        form.fields['semester'].queryset = semesters

        # Filter branches based on the department of the logged-in HOD
        branches = Branch.objects.filter(department=hod.department)

        # Pass the filtered queryset to the form
        form.fields['branches'].queryset = branches

        

    return render(request, 'hod/add_subject.html', {'form': form})




def edit_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            return redirect('hod-dashboard')  # Redirect to the dashboard after saving
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'hod/edit_subject.html', {'form': form, 'subject': subject})

from django.shortcuts import render
from .models import Semester

def hod_semesters(request):
    if request.user.role == 'hod':
        # Retrieve semesters belonging to the HOD's department
        hod_department = request.user.hod.department
        semesters = Semester.objects.filter(branch__department=hod_department)
        
        context = {
            'semesters': semesters,
        }

        return render(request, 'hod/semesters.html', context)
    
    # If the user is not an HOD, handle accordingly
    return HttpResponseForbidden("You don't have permission to access this page.")




@login_required(login_url='login')
def edit_semester(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    if request.method == 'POST':
        form = SemesterForm(request.POST, instance=semester)
        if form.is_valid():
            form.save()
            return redirect('hod-dashboard')
    else:
        form = SemesterForm(instance=semester)
    return render(request, 'hod/edit_semester.html', {'form': form})

def LogoutPage(request):
    logout(request)
    return redirect('login')






# def is_teacher(user):
#     return user.role == 'teacher'
# @user_passes_test(is_teacher)


# @permission_required('registration_app.can_mark_attendance')
@login_required(login_url='login')
@transaction.atomic
def mark_attendance(request, subject_id):
    try:
        # Check the user's role and get the related teacher
        if request.user.role == 'teacher':
            teacher = request.user  # Assign the user object directly
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    except CustomUser.DoesNotExist:
        return HttpResponseForbidden("You don't have permission to access this page.")

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        date_from = datetime.strptime(request.POST.get('date_from'), '%Y-%m-%d').date()
        date_to = datetime.strptime(request.POST.get('date_to'), '%Y-%m-%d').date()

        # Ensure attendance can only be marked for the present date range
        current_date = datetime.today().date()
        if date_from != current_date or date_to != current_date:
            messages.error(request, "You can only mark attendance for the present date range.")
            return redirect('mark_attendance', subject_id=subject_id)

        # Check if attendance for the subject and date range already exists
        if Attendance.objects.filter(subject=subject, date__range=[date_from, date_to]).exists():
            messages.error(request, "Attendance for this subject and date range already exists.")
            return redirect('mark_attendance', subject_id=subject_id)

        # Save the Attendance instances before creating AttendanceReport instances
        attendance = []
        for single_date in daterange(date_from, date_to):
            attendance_instance = Attendance.objects.create(subject=subject, date=single_date, submitted_by=request.user)
            attendance.append(attendance_instance)

        students = subject.students.all()

        for att in attendance:
            for student in students:
                status = request.POST.get(f'student_{student.id}', False)
                attendance_report, created = AttendanceReport.objects.get_or_create(student=student, attendance=att)
                attendance_report.status = status
                attendance_report.save()

        subject.total_classes = Attendance.objects.filter(subject=subject).count()
        subject.save()

        messages.success(request, "Attendance marked successfully.")
        return redirect('mark_attendance', subject_id=subject_id)

    # Add messages to the request context before rendering the template
    messages_to_display = messages.get_messages(request)

    return render(request, 'mark_attendance.html', {'subject': subject, 'messages': messages_to_display})

def daterange(start_date, end_date):
    """
    Generate a range of dates between start_date and end_date.
    """
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)



@login_required
def student_attendance(request):
    # Get the current student
    student = request.user.student

    # Get all subjects associated with the student
    subjects = student.subjects.all()

    # Fetch attendance reports for all subjects
    attendance_reports = AttendanceReport.objects.filter(student=student)

    return render(request, 'student_attendance.html', {'subjects': subjects, 'attendance_reports': attendance_reports})


def subject_attendance(request, subject_id):
    # Get the subject object
    subject = get_object_or_404(Subject, id=subject_id)
    # Get the current student
    student = request.user.student
    # Fetch attendance reports for the specified subject and student
    attendance_reports = AttendanceReport.objects.filter(student=student, attendance__subject=subject)
    return render(request, 'subject_attendance.html', {'subject': subject, 'attendance_reports': attendance_reports})



def create_subject(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirect to a success page or any other appropriate action
            return redirect('success_page')
    else:
        form = SubjectForm()
    return render(request, 'subject_form.html', {'form': form})



def success_page(request):
    # You can add any custom logic here
    return render(request, 'success_page.html')


@login_required(login_url='login')
def teacher_dashboard(request):
    if request.user.role == 'teacher':
        teacher = Teacher.objects.get(user=request.user)

        # Retrieve subjects assigned to the teacher
        subjects = teacher.subjects.all()
        announcements = Announcement.objects.all().order_by('-posted_on')[:3]

        # Fetch leave data
        leaves = Leave.objects.filter(teacher=teacher).order_by('-applied_on')

        context = {
            'teacher': teacher,
            'subjects': subjects,
            'announcements': announcements,
            'leaves': leaves,
        }

        return render(request, 'teacher_dashboard.html', context)

    return HttpResponseForbidden("You don't have permission to access this page.")


@login_required
def teacher_leave_status(request):
    """Display the teacher's leave requests and their status."""
    leaves = Leave.objects.filter(teacher__user=request.user).order_by('-start_date')
    return render(request, 'teacher_leave_status.html', {'leaves': leaves})


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Leave
from django.contrib import messages
from django.http import HttpResponseForbidden
from datetime import datetime

@login_required(login_url='login')
def apply_leave(request):
    if request.user.role == 'teacher':
        if request.method == 'POST':
            leave_type = request.POST.get('leave_type')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            reason = request.POST.get('reason')

            # Check if teacher has already applied for 2 leaves in the current month
            current_month = datetime.now().month
            current_year = datetime.now().year

            # Count the number of leaves already applied by the teacher in the current month and year
            leave_count = Leave.objects.filter(
                teacher=request.user.teacher,
                start_date__month=current_month,
                start_date__year=current_year
            ).count()

            # If teacher has already applied for two leaves, show an error message
            if leave_count >= 4:
                messages.error(request, "You can only apply for two leaves per month.")
                return render(request, 'apply_leave.html')

            # Save leave request if limit not reached
            Leave.objects.create(
                teacher=request.user.teacher,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                reason=reason,
                status="Pending"
            )

            messages.success(request, "Your leave request has been submitted successfully.")
            return render(request,'apply_leave.html')

        return render(request, 'apply_leave.html')

    return HttpResponseForbidden("You don't have permission to access this page.")




def subject_details(request, subject_id):
    subject = Subject.objects.get(id=subject_id)
    students = subject.students.all()
    return render(request, 'subject_details.html', {'subject': subject, 'students': students})


def student_details(request, student_id):
    # Get the Student object by student_id, assuming you have a unique identifier.
    student = get_object_or_404(Student, pk=student_id)

    student_name = student.name
    student_contact = student.contact
    student_address = student.address
    
    # You can now use these variables in your template or return them in an HttpResponse.
    
    return render(request, 'student_details.html', {
        'student_name': student_name,
        'student_contact': student_contact,
        'student_address': student_address,
        'student_user_email': student.user.email,  # Pass the email to the template
    })
    
    # Access the email attribute through the student instance.
    student_user_email = student.user.email

    
    
    # You can now use student_user_email in your template or return it in an HttpResponse.
    
    return render(request, 'student_details.html', {'student_user_email': student_user_email})

@login_required(login_url='login')

def edit_profile(request):
    if request.user.is_authenticated and hasattr(request.user, 'student'):
        student = request.user.student

        if request.method == 'POST':
            edit_profile_form = StudentEditForm(request.POST, request.FILES, instance=student)
            if edit_profile_form.is_valid():
                student = edit_profile_form.save(commit=False)
                student.profile_picture = request.FILES.get('profile_picture')  # Handle profile picture upload
                student.save()
                return redirect('student_dashboard')

        else:
            edit_profile_form = StudentEditForm(instance=student)

        return render(request, 'edit_profile.html', {'edit_profile_form': edit_profile_form})

    return redirect('login')


def select_subject_to_mark_attendance(request):
    subjects = Subject.objects.all()  # Query all subjects
    return render(request, 'select_subject_to_mark_attendance.html', {'subjects': subjects})




@login_required
def all_subjects(request):
    # Get the currently logged-in user (CustomUser instance)
    user = request.user
    
    try:
        # Retrieve the associated Teacher instance
        teacher = Teacher.objects.get(user=user)
        
        # Get the subjects taught by the teacher
        subjects = teacher.subjects.all()
        
        return render(request, 'all_subjects.html', {'subjects': subjects})
    
    except Teacher.DoesNotExist:
        # Handle the case where the logged-in user is not a teacher
        return HttpResponseForbidden("You don't have permission to access this page.")




def download_attendance_report(request, subject_id, date_from, date_to):
    # Parse the date_from and date_to strings into datetime objects
    date_from = dt.strptime(date_from, '%Y-%m-%d').date()
    date_to = dt.strptime(date_to, '%Y-%m-%d').date()

    # Fetch subject and attendance records within the specified date range
    subject = get_object_or_404(Subject, id=subject_id)
    attendance_records = Attendance.objects.filter(subject=subject, date__range=[date_from, date_to])

    # Create a new Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Set column headers
    ws.append(['Date', 'Roll Number', 'Student ID', 'Branch', 'Semester','Subject', 'Student Name', 'Attendance Status'])

    # Add attendance records to the worksheet
    for record in attendance_records:
        attendance_reports = AttendanceReport.objects.filter(attendance=record)
        for attendance_report in attendance_reports:
            # Check if the student has a branch
            if attendance_report.student.branch:
                branch_name = attendance_report.student.branch.name
            else:
                branch_name = "N/A"

            semester_name = attendance_report.student.semester.name
            subject_name = record.subject.name
            student_name = attendance_report.student.name

            # Customize the report format as needed
            status_display = "Present" if attendance_report.status else "Absent"
            ws.append([record.date, attendance_report.student.roll_number, 
                       attendance_report.student.student_id, branch_name, semester_name,
                       subject_name, student_name, status_display])

    # Set column widths and alignment
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Create a response object for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{date_from}_{date_to}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response





from django.shortcuts import render, redirect
from .forms import MarksUploadForm
from .models import Teacher

def upload_marks(request):
    if request.method == 'POST':
        form = MarksUploadForm(request.POST, request.FILES)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            excel_file = request.FILES['excel_file']
            # Your code to process the Excel file and save marks
            return redirect('teacher_dashboard')  # Redirect to the teacher dashboard after successful upload
    else:
        if request.user.is_authenticated and request.user.role == 'teacher':
            try:
                teacher = Teacher.objects.get(user=request.user)
                subjects = teacher.subjects.all()
                form = MarksUploadForm(initial={'subject': subjects})
            except Teacher.DoesNotExist:
                # Handle the case where no Teacher object exists for the user
                form = MarksUploadForm()
        else:
            form = MarksUploadForm()
    return render(request, 'upload_marks.html', {'form': form})







def view_marks(request):
    # Get distinct exam types from the Marks model
    exam_types = Marks.objects.values_list('exam_type', flat=True).distinct()

    # Get the selected exam type from the request GET parameters
    selected_exam_type = request.GET.get('exam_type')

    # Check if the user is authenticated and is a student
    if request.user.is_authenticated and hasattr(request.user, 'student'):
        # Get student details
        student_roll_number = request.user.student.roll_number
        student_semester = request.user.student.semester.name
        student_branch = request.user.student.branch.name

        # Filter marks based on selected exam type, student roll number, semester, and branch
        if selected_exam_type:
            marks = Marks.objects.filter(exam_type=selected_exam_type,
                                          roll_number=student_roll_number,
                                          semester=student_semester,
                                          branch=student_branch)
        else:
            marks = None

        context = {
            'exam_types': exam_types,
            'selected_exam_type': selected_exam_type,
            'marks': marks
        }
        return render(request, 'view_marks.html', context)
    else:
        # Handle unauthorized access or non-student user
        return HttpResponse("Unauthorized access or user is not a student.")

def parse_excel_file(file):
    marks_data = []
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file)
    # Assuming the first sheet is the one containing marks data
    sheet = workbook.active
    # Iterate over rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Assuming the columns are in the order: Student_id, Roll Number, Semester, Branch, Subject, Exam Type, Student Name, Marks
        marks_obtained = None
        try:
            marks_obtained = Decimal(row[7])
        except InvalidOperation:
            try:
                marks_obtained = Decimal(str(row[7]))
            except InvalidOperation:
                pass
        marks_data.append({ 
            'student_id': row[0],
            'roll_number': row[1],
            'semester': row[2],
            'branch': row[3],
            'subject': row[4],
            'exam_type': row[5],
            'student_name': row[6],
            'marks_obtained': marks_obtained
        })
    return marks_data


def save_marks_data(subject, marks_data):
    for data in marks_data:
        marks = Marks.objects.create(
            subject=subject,
            student_id=data['student_id'],
            roll_number=data['roll_number'],
            semester=data['semester'],
            branch=data['branch'],
            student_name=data['student_name'],
            marks_obtained=data['marks_obtained'],
            exam_type=data['exam_type']
        )

def student_announcements(request):
    announcements = Announcement.objects.all()  # Get all announcements
    return render(request, 'student_announcements.html', {'announcements': announcements})

def teacher_announcements(request):
    announcements = Announcement.objects.all()  # Get all announcements
    return render(request, 'teacher_announcements.html', {'announcements': announcements})
 







from django.shortcuts import render, get_object_or_404
from .models import Course, Branch, Semester
from django.contrib.auth.decorators import login_required

@login_required(login_url='login')
def account_section(request):
    courses = Course.objects.all()
    return render(request, 'account_section.html', {'courses': courses})

@login_required(login_url='login')
def branch_list(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    branches = Branch.objects.filter(course=course)
    return render(request, 'branch_list.html', {'course': course, 'branches': branches})

@login_required(login_url='login')
def semester_list(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    semesters = Semester.objects.filter(branch=branch)
    return render(request, 'semester_list.html', {'branch': branch, 'semesters': semesters})

from .models import HostelFees

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import Student, TransportFees, HostelFees

@login_required(login_url='login')
def view_fees(request):
    try:
        student = request.user.student
        fee_structure = student.fee_structure  # Get linked FinancialFees
        miscellaneous_fees = student.miscellaneous_fees
        fees_fine = student.fees_fine

        if not fee_structure:
            return HttpResponseForbidden("Fee structure is not available for this student.")
        
        total_fees = fee_structure.registration_fees + fee_structure.academic_fees
        paid_amount = student.get_paid_amount()
        hostel_fees = None
        transport_fees = None  # Initialize transport_fees

        # Ensure the student has a hostel room assigned
        if student.hostel_room and student.hostel_room.hostel:
            hostel_fees_obj = HostelFees.objects.filter(hostel_details=student.hostel_room).first()
            if hostel_fees_obj:
                hostel_fees = hostel_fees_obj.fee_amount
                total_fees += hostel_fees

        # Retrieve transport fees
        if hasattr(student, 'transport_fees') and student.transport_fees:
            transport_fees = student.transport_fees.fees_amount
            total_fees += transport_fees  # Add transport fees to total

        total_fees += miscellaneous_fees + fees_fine
        due_amount = total_fees - paid_amount

        return render(
            request,
            "fees.html",
            {
                "fee_structure": fee_structure,
                "miscellaneous_fees": miscellaneous_fees,
                "fees_fine": fees_fine,
                "total_fees": total_fees,
                "paid_amount": paid_amount,
                "due_amount": due_amount,
                "hostel_fees": hostel_fees,
                "transport_fees": transport_fees,  # Send transport fees to the template
            },
        )

    except Student.DoesNotExist:
        return HttpResponseForbidden("You don't have permission to access this page.")





from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import Student, Payment, TransportFees, HostelFees
from .forms import PaymentForm

@login_required(login_url='login')
def pay_fees(request):
    try:
        student = request.user.student
        fee_structure = student.fee_structure

        if not fee_structure:
            messages.error(request, "Fee structure is not assigned to you. Please contact the administration.")
            return redirect('student_dashboard')

        # Base fees
        total_fees = fee_structure.registration_fees + fee_structure.academic_fees

        # Add miscellaneous fees and fine
        total_fees += student.miscellaneous_fees + student.fees_fine

        # Calculate paid amount
        paid_amount = Payment.objects.filter(student=student, status='Completed').aggregate(total=Sum('amount'))['total'] or 0

        # Debug transport fees
        transport_fees = None
        print(f"Student Transport Fees ID: {student.transport_fees_id}")  # Debugging
        
        if student.transport_fees_id:
            transport_fees_obj = TransportFees.objects.filter(id=student.transport_fees_id).first()
            if transport_fees_obj:
                transport_fees = transport_fees_obj.fees_amount
                total_fees += transport_fees
                print(f"Transport Fees Amount: {transport_fees}")  # Debugging

        # Add hostel fees (if applicable)
        hostel_fees = None
        if student.hostel_room:
            hostel_fees_obj = HostelFees.objects.filter(hostel_details=student.hostel_room).first()
            if hostel_fees_obj:
                hostel_fees = hostel_fees_obj.fee_amount
                total_fees += hostel_fees

        # Calculate due amount
        due_amount = max(total_fees - paid_amount, 0)

        if due_amount == 0:
            messages.success(request, "You have already completed your payment.")
            return redirect('student_dashboard')

        if request.method == "POST":
            form = PaymentForm(request.POST)
            if form.is_valid():
                payment = form.save(commit=False)
                payment.student = student
                payment.status = 'Pending'  # Default to pending before confirmation
                payment.save()
                messages.success(request, "Payment initiated. Await confirmation.")
                return redirect('student_dashboard')
        else:
            form = PaymentForm()

        return render(request, 'pay_fees.html', {
            'student': student,
            'form': form,
            'remaining_amount': due_amount,
            'miscellaneous_fees': student.miscellaneous_fees,
            'fees_fine': student.fees_fine,
            'due_amount': due_amount,
            'hostel_fees': hostel_fees,
            'transport_fees': transport_fees  # Ensure transport fees are passed
        })

    except Student.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect('login')



 









from django.db.models import Sum
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Course, Branch, Semester, Student, FeeStructure
from .forms import AssignFeeStructureForm, PaymentForm

@login_required(login_url='login')
def semester_student_list(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    students = Student.objects.filter(semester=semester)

    for student in students:
        total_paid = student.payment_set.aggregate(total_paid=Sum('amount'))['total_paid'] or 0
        fee_structure = student.fee_structure
        if fee_structure:
            total_amount = fee_structure.amount
            remaining_amount = max(total_amount - total_paid, 0)
            student.total_paid = total_paid
            student.remaining_amount = remaining_amount
        else:
            student.total_paid = total_paid
            student.remaining_amount = 0

    return render(request, 'semester_student_list.html', {'semester': semester, 'students': students})

@login_required
def assign_fee_structure(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        form = AssignFeeStructureForm(request.POST)
        if form.is_valid():
            student.fee_structure = form.cleaned_data['fee_structure']
            student.save()
            messages.success(request, 'Fee structure assigned successfully.')
            return redirect('semester_student_list', semester_id=student.semester.id)
        else:
            messages.error(request, 'Invalid data. Please try again.')
    else:
        form = AssignFeeStructureForm()

    return render(request, 'assign_fee_structure.html', {'form': form, 'student': student})

# @login_required(login_url='login')
# def pay_fees(request, student_id):
#     student = get_object_or_404(Student, id=student_id)  # Get the student by ID

#     try:
#         fee_structure = student.fee_structure
#         if not fee_structure:
#             raise FeeStructure.DoesNotExist
#     except FeeStructure.DoesNotExist:
#         return HttpResponseForbidden("Fee structure is not available for this student.")

#     if request.method == 'POST':
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             payment = form.save(commit=False)
#             payment.student = student
#             payment.save()
#             messages.success(request, 'Payment recorded successfully.')
#             return redirect('semester_student_list', semester_id=student.semester.id)
#     else:
#         form = PaymentForm()

#     total_paid = student.payment_set.aggregate(total_paid=Sum('amount'))['total_paid'] or 0
#     remaining_amount = fee_structure.amount - total_paid

#     return render(request, 'pay_fees.html', {'form': form, 'remaining_amount': remaining_amount, 'student': student})


@login_required(login_url='login')
def student_section_dashboard(request):
    if request.user.role != 'ss':
        return HttpResponse("Unauthorized", status=401)
    students = Student.objects.all()  # Fetch all students
    return render(request, 'student_section_dashboard.html', {'students': students})


@login_required(login_url='login')
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('account_section')  # Redirect to account section after saving
    else:
        form = StudentForm()
    
    return render(request, 'fill_details.html', {'form': form})
 

from .forms import AssignFeeStructureForm





from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Course, Branch, Semester, Student

@login_required
def update_student_view(request):
    courses = Course.objects.all()
    branches = Branch.objects.all()
    semesters = Semester.objects.all()

    students = None
    if request.method == 'POST':
        selected_course = request.POST.get('course')
        selected_branch = request.POST.get('branch')
        selected_semester = request.POST.get('semester')

        students = Student.objects.filter(
            course_id=selected_course,
            branch_id=selected_branch,
            semester_id=selected_semester
        )

    context = {
        'courses': courses,
        'branches': branches,
        'semesters': semesters,
        'students': students,
    }
    return render(request, 'update_student.html', context)

@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect(reverse('update_student'))
    else:
        form = StudentForm(instance=student)
    
    context = {
        'form': form,
        'student': student
    }
    return render(request, 'edit_student.html', context)








@login_required
def assign_fee_structure(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        form = AssignFeeStructureForm(request.POST)
        if form.is_valid():
            student.fee_structure = form.cleaned_data['fee_structure']
            student.save()
            messages.success(request, 'Fee structure assigned successfully.')
            return redirect('account_section')
        else:
            messages.error(request, 'Invalid data. Please try again.')
    else:
        form = AssignFeeStructureForm()

    return render(request, 'assign_fee_structure.html', {'form': form, 'student': student})


@login_required(login_url='login')
def add_student_certificate(request):
    if request.method == 'POST':
        form = StudentCertificateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('student_section_dashboard')  # Redirect to student dashboard after saving
    else:
        form = StudentCertificateForm()
    
    return render(request, 'add_student_certificate.html', {'form': form})



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Student, StudentCertificate
from .forms import StudentCertificateForm
from django.contrib.auth.decorators import login_required

@login_required
def student_documents_list(request):
    students_with_documents = Student.objects.filter(studentcertificate__isnull=False).distinct()
    return render(request, 'student_documents_list.html', {'students': students_with_documents})

@login_required
def student_documents_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    certificates = student.studentcertificate_set.all()

    if request.method == 'POST':
        form = StudentCertificateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Certificate uploaded successfully.')
            return redirect('student_documents_detail', student_id=student.id)
    else:
        form = StudentCertificateForm(initial={'student': student})

    return render(request, 'student_documents_detail.html', {
        'student': student,
        'certificates': certificates,
        'form': form
    })


@login_required
def delete_certificate(request, certificate_id):
    certificate = get_object_or_404(StudentCertificate, id=certificate_id)
    student_id = certificate.student.id
    certificate.delete()
    messages.success(request, 'Certificate deleted successfully.')
    return redirect('student_documents_detail', student_id=student_id)

        


 


def select_course(request):
    courses = Course.objects.all()
    return render(request, 'select_course_student.html', {'courses': courses})

def select_program(request, course_id):
    programs = Branch.objects.filter(course_id=course_id)
    return render(request, 'select_program_student.html', {'programs': programs})

def select_semester(request, program_id):
    semesters = Semester.objects.filter(branch_id=program_id)
    return render(request, 'select_semester_student.html', {'semesters': semesters})

def list_students(request, semester_id):
    students = Student.objects.filter(semester_id=semester_id)
    return render(request, 'list_students_student.html', {'students': students})
 




# HR views

from django.shortcuts import render, redirect
from .forms import AddNonTeachingStaffForm

def hr_admin_dashboard(request):
    return render(request, 'hr_admin/base_hr_admin.html')

from django.shortcuts import render, get_object_or_404
from .models import Teacher
from .forms import TeacherForm
 
def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            teacher = form.save(commit=False)
            teacher.user = get_object_or_404(CustomUser, id=request.POST.get('user_id'))
            teacher.save()
            return redirect('hr_admin_dashboard')  # Redirect to HR Admin Dashboard or wherever needed
    else:
        form = TeacherForm()
        # Filter to get only users with the role 'teacher'
        users = CustomUser.objects.filter(role='teacher')
        context = {
            'form': form,
            'users': users,
        }
    return render(request, 'hr_admin/add_teacher.html', context)



def add_non_teaching_staff(request):
    if request.method == 'POST':
        form = AddNonTeachingStaffForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hr_admin_dashboard')
    else:
        form = AddNonTeachingStaffForm()

    return render(request, 'hr_admin/add_non_teaching_staff.html', {'form': form})



# FEES STRUCTURE 

def get_student_fees(student):
    try:
        fees = FinancialFees.objects.get(school=student.department.school, program=student.course)
        total_fees = (
            fees.registration_fees +
            fees.academic_fees +
            (fees.hostel_fees.amount if fees.hostel_fees else 0) + 
            fees.miscellaneous_fees +
            fees.fees_fine
        )
        return total_fees
    except FinancialFees.DoesNotExist:
        return "Fees structure not defined for this student"


 

